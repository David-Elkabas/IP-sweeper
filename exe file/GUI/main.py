import subprocess
import concurrent.futures
import platform
import openpyxl
from datetime import datetime
from Classes import IP_address

def read_old_data():
    path = "output_data.xlsx"
    print("start fetching data")
    # To open the workbook, workbook object is created
    workbook_obj = openpyxl.load_workbook(path)

    return get_data_from_file(workbook_obj)


def get_date():
    now = datetime.now()  # current date and time
    date_time = now.strftime("%d/%m/%Y, %H:%M:%S")

    return date_time


def ping_ip(current_ip_address):
    try:
        output = subprocess.check_output("ping -{} 1 {}".format('n' if platform.system().lower(
        ) == "windows" else 'c', current_ip_address), shell=True, universal_newlines=True)
        if 'unreachable' in output:
            return "not work"
        else:
            return "work"
    except Exception:
        return "not work"

def get_data_from_file(workbook_obj):
    # Give the location of the file

    # Get workbook active sheet object from the active attribute
    sheet_obj = workbook_obj.active

    # Note: The first row or column integer is 1, not 0.

    all_rows = []
    only_names = []
    only_ip = []
    only_status = []
    only_modify = []

    for row in sheet_obj.rows:
        current_row = []
        cell_count = 1
        for cell in row:
            current_row.append(cell.value)
            if cell_count == 1:
                only_names.append(cell.value)
            if cell_count == 2:
                only_ip.append(cell.value)
            if cell_count == 3:
                only_status.append(cell.value)
            if cell_count == 4:
                only_modify.append(cell.value)
            cell_count += 1
        all_rows.append(current_row)

    only_names.pop(0)
    only_ip.pop(0)
    only_status.pop(0)
    only_modify.pop(0)

    all_addresses = {}
    for name, ip, status, modify in zip(only_names,only_ip,only_status,only_modify):
         all_addresses[name] = IP_address(name, ip, status, modify)

    return all_addresses


def add_data_to_file(path, address_name, address_ip):

    workbook_obj = openpyxl.load_workbook(path)
    sheet_obj = workbook_obj.active

    max_row = sheet_obj.max_row
    name_cell = 'A' + str(max_row+1)
    ip_cell = 'B' + str(max_row+1)

    sheet_obj[name_cell]=address_name
    sheet_obj[ip_cell]=address_ip

    workbook_obj.save(path)


def edit_file(workbook_obj, address_name, status, time):
    # # Give the location of the file
    # path = "C:\\Users\\David\\Desktop\\code projects\\pingger\\source\\data.xlsx"
    #
    # # To open the workbook, workbook object is created
    # workbook_obj = openpyxl.load_workbook(path)

    # Get workbook active sheet object from the active attribute
    sheet_obj = workbook_obj.active
    cell_positions = ''
    for row in sheet_obj.rows:
        element = row[0]
        if element.value == address_name:
            cell_col_position = 'C'
            cell_row_position = element.row
            cell_positions = cell_col_position + str(cell_row_position)
            time_cell_positions = 'D' + str(cell_row_position)

    if status == 'work':
        sheet_obj[cell_positions] = "work"
        sheet_obj[cell_positions].fill = fill = openpyxl.styles.GradientFill(stop=("80ff72", "7ee8fa"))
        # print( sheet_obj[cell_positions].value)

    else:
        sheet_obj[cell_positions] = "not work"
        # print(sheet_obj[cell_positions].value)
        sheet_obj[cell_positions].fill = fill = openpyxl.styles.GradientFill(stop=("fc9842", "fe5f75"))

    sheet_obj[time_cell_positions] = time


def update_data(list_of_all_addresses, key, status, time):
    list_of_all_addresses[key].set_status(status)
    list_of_all_addresses[key].set_modify(time)
    return list_of_all_addresses


def scanning_all(dict_of_all_addresses,workbook_obj):
    list_address = []
    for key in dict_of_all_addresses:
        list_address.append(dict_of_all_addresses[key].ip)

    with concurrent.futures.ThreadPoolExecutor() as executor:
        results = executor.map(ping_ip, list_address)

    i = 0
    results = list(results)
    for key in dict_of_all_addresses:
        dict_of_all_addresses[key].status = results[i]
        time = get_date()
        dict_of_all_addresses[key].modify = time
        i += 1

    for key in dict_of_all_addresses:
        edit_file(workbook_obj, key, dict_of_all_addresses[key].status, dict_of_all_addresses[key].modify)


def main_algo():

    path = "first_data.xlsx"

    print("start fetching data")
    # To open the workbook, workbook object is created
    workbook_obj = openpyxl.load_workbook(path)

    list_of_all_addresses = get_data_from_file(workbook_obj)

    print("finish fetching data")

    # IP_address.print_all_addresses(list_of_all_addresses)
    print("start sending ping to all the Ip addresses")

    test_mode = False

    if (test_mode):
        for dict_key in list_of_all_addresses:
        # ------- for testing the GUI only ---------
            time = get_date()
            status = "dont work"
            edit_file(workbook_obj, dict_key, status, time)
            list_of_all_addresses = update_data(list_of_all_addresses, dict_key, status, time)
        # ------------------------------------------
    else:
        scanning_all(list_of_all_addresses, workbook_obj)


    print("finish sending all the pings")
    # IP_address.print_all_addresses(list_of_all_addresses)
    # save the file after editing
    workbook_obj.save('output_data.xlsx')

    print("saved all new data to output_data.xlsx")
    return list_of_all_addresses
